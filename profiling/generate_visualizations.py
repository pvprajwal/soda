#!/usr/bin/env python3
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_FILE = os.path.join(PROJECT_DIR, "profiling_results.csv")
GRAPH_ALL = os.path.join(PROJECT_DIR, "performance_all_metrics.png")
GRAPH_CPU_GPU = os.path.join(PROJECT_DIR, "performance_cpu_vs_gpu.png")
EXCEL_FILE = os.path.join(PROJECT_DIR, "profiling_results_presentation_v3.xlsx")

# Read CSV
df = pd.read_csv(CSV_FILE)

print("Creating graphs...")

# Create figure with two subplots
fig, axes = plt.subplots(1, 2, figsize=(16, 6))

# Graph 1: All three metrics
ax1 = axes[0]
ax1.plot(df['LIST_SIZE'], df['CPU_Time_ms'], marker='o', linewidth=2, label='CPU Time', color='#1f77b4')
ax1.plot(df['LIST_SIZE'], df['GPU_Time_ms'], marker='s', linewidth=2, label='GPU Kernel Time', color='#ff7f0e')
ax1.plot(df['LIST_SIZE'], df['Total_GPU_Execution_ms'], marker='^', linewidth=2, label='Total GPU Execution', color='#2ca02c')
ax1.set_xlabel('List Size', fontsize=12, fontweight='bold')
ax1.set_ylabel('Time (ms)', fontsize=12, fontweight='bold')
ax1.set_title('Performance Comparison: All Metrics', fontsize=14, fontweight='bold')
ax1.set_xscale('log')
ax1.set_yscale('log')
ax1.grid(True, which='both', alpha=0.3)
ax1.legend(fontsize=11, loc='best')

# Graph 2: CPU vs GPU only
ax2 = axes[1]
ax2.plot(df['LIST_SIZE'], df['CPU_Time_ms'], marker='o', linewidth=2.5, label='CPU Time', color='#d62728', markersize=7)
ax2.plot(df['LIST_SIZE'], df['GPU_Time_ms'], marker='s', linewidth=2.5, label='GPU Kernel Time', color='#17becf', markersize=7)
ax2.set_xlabel('List Size', fontsize=12, fontweight='bold')
ax2.set_ylabel('Time (ms)', fontsize=12, fontweight='bold')
ax2.set_title('CPU vs GPU Kernel Execution Time', fontsize=14, fontweight='bold')
ax2.set_xscale('log')
ax2.set_yscale('log')
ax2.grid(True, which='both', alpha=0.3)
ax2.legend(fontsize=11, loc='best')

plt.tight_layout()
plt.savefig(GRAPH_ALL, dpi=300, bbox_inches='tight')
print(f"✓ Saved: {GRAPH_ALL}")

# Also save individual graphs for clarity
fig1, ax = plt.subplots(figsize=(12, 7))
ax.plot(df['LIST_SIZE'], df['CPU_Time_ms'], marker='o', linewidth=2, label='CPU Time', color='#1f77b4', markersize=6)
ax.plot(df['LIST_SIZE'], df['GPU_Time_ms'], marker='s', linewidth=2, label='GPU Kernel Time', color='#ff7f0e', markersize=6)
ax.plot(df['LIST_SIZE'], df['Total_GPU_Execution_ms'], marker='^', linewidth=2, label='Total GPU Execution', color='#2ca02c', markersize=6)
ax.set_xlabel('List Size', fontsize=13, fontweight='bold')
ax.set_ylabel('Time (ms)', fontsize=13, fontweight='bold')
ax.set_title('Performance Comparison: All Metrics', fontsize=15, fontweight='bold')
ax.set_xscale('log')
ax.set_yscale('log')
ax.grid(True, which='both', alpha=0.3)
ax.legend(fontsize=12, loc='best')
plt.tight_layout()
plt.savefig(GRAPH_ALL, dpi=300, bbox_inches='tight')
plt.close()

fig2, ax = plt.subplots(figsize=(12, 7))
ax.plot(df['LIST_SIZE'], df['CPU_Time_ms'], marker='o', linewidth=2.5, label='CPU Time', color='#d62728', markersize=8)
ax.plot(df['LIST_SIZE'], df['GPU_Time_ms'], marker='s', linewidth=2.5, label='GPU Kernel Time', color='#17becf', markersize=8)
ax.set_xlabel('List Size', fontsize=13, fontweight='bold')
ax.set_ylabel('Time (ms)', fontsize=13, fontweight='bold')
ax.set_title('CPU vs GPU Kernel Execution Time', fontsize=15, fontweight='bold')
ax.set_xscale('log')
ax.set_yscale('log')
ax.grid(True, which='both', alpha=0.3)
ax.legend(fontsize=12, loc='best')
plt.tight_layout()
plt.savefig(GRAPH_CPU_GPU, dpi=300, bbox_inches='tight')
plt.close()

print(f"✓ Saved: {GRAPH_CPU_GPU}")

# Export half the results (alternate rows) to Excel with styling
print("\nCreating styled Excel file...")
df_export = df.iloc[::2].reset_index(drop=True)  # Skip alternate rows

# Also include the 1024000 row specifically
row_1024000 = df[df['LIST_SIZE'] == 1024000]
if not row_1024000.empty:
    df_export = pd.concat([df_export, row_1024000], ignore_index=True).drop_duplicates().sort_values('LIST_SIZE').reset_index(drop=True)

wb = Workbook()
ws = wb.active
ws.title = "Profiling Results"

# Header styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data styling
data_fill_light = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
data_fill_dark = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
data_font = Font(size=11)
data_alignment = Alignment(horizontal="center", vertical="center")

# Border styling
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add headers
headers = ['LIST_SIZE', 'CPU_Time_ms', 'GPU_Time_ms', 'Total_GPU_Execution_ms']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Add data rows
for row_num, (idx, row) in enumerate(df_export.iterrows(), 2):
    # Alternate row colors
    fill_color = data_fill_light if row_num % 2 == 0 else data_fill_dark
    
    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        value = row[col_name]
        
        # Format numbers
        if col_num == 1:  # LIST_SIZE
            cell.value = int(value)
            cell.number_format = '#,##0'
        else:  # Time values
            cell.value = round(value, 4)
            cell.number_format = '0.0000'
        
        cell.fill = fill_color
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 22

# Freeze header row
ws.freeze_panes = "A2"

wb.save(EXCEL_FILE)
print(f"✓ Saved: {EXCEL_FILE}")
print(f"  - Exported {len(df_export)} rows (every other row from original {len(df)} rows)")

print("\n✅ All outputs created successfully!")
print(f"   - {GRAPH_ALL}")
print(f"   - {GRAPH_CPU_GPU}")
print(f"   - {EXCEL_FILE}")
